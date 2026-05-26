"""Bulk-upload Meta ads from a CSV or XLSX template.

All entities are created with status=PAUSED so nothing spends money until you
review and activate in Ads Manager.

Env vars required:
  META_ACCESS_TOKEN     long-lived access token with ads_management scope
  META_AD_ACCOUNT_ID    e.g. act_1234567890

Usage:
  python bulk_upload.py template.csv
  python bulk_upload.py template.xlsx
  python bulk_upload.py template.csv --dry-run
"""

import argparse
import csv
import json
import os
import re
import sys
import urllib.request
from collections import defaultdict

from facebook_business.adobjects.ad import Ad
from facebook_business.adobjects.adaccount import AdAccount
from facebook_business.adobjects.adcreative import AdCreative
from facebook_business.adobjects.adset import AdSet
from facebook_business.adobjects.campaign import Campaign
from facebook_business.api import FacebookAdsApi

PAUSED = "PAUSED"

# Meta amounts are sent in the ad account's atomic currency unit
# (e.g. cents for USD, won for KRW). These three lists categorize ISO
# 4217 currency codes by how many minor units they have. Anything not
# listed defaults to 2 decimals (multiplier=100), which matches USD/EUR
# and the majority of other currencies.
_ZERO_DECIMAL_CURRENCIES = {
    "BIF", "CLP", "COP", "DJF", "GNF", "ISK", "JPY", "KMF", "KRW",
    "PYG", "RWF", "UGX", "VND", "VUV", "XAF", "XOF", "XPF",
}
_THREE_DECIMAL_CURRENCIES = {"BHD", "IQD", "JOD", "KWD", "LYD", "OMR", "TND"}


def _currency_multiplier(currency):
    """Atomic-unit multiplier for the ad account's currency. USD 50.00
    becomes 5000; KRW 50000 becomes 50000; BHD 50.000 becomes 50000."""
    c = (currency or "").upper()
    if c in _ZERO_DECIMAL_CURRENCIES:
        return 1
    if c in _THREE_DECIMAL_CURRENCIES:
        return 1000
    return 100


def _money(row, key):
    """Read a money cell, accepting the legacy `_usd` suffix as a
    fallback so spreadsheets filled before the currency rename still
    work."""
    val = _get(row, key)
    if val:
        return val
    return _get(row, key + "_usd")


def _parse_amount(value):
    """Parse a money string into a float, tolerating thousand-separator
    commas (20,000) and surrounding currency symbols / whitespace."""
    if value is None or value == "":
        return 0.0
    s = str(value).strip().replace(",", "").replace("$", "").replace(" ", "")
    return float(s)


def convert_drive_url(url):
    """Rewrite Google Drive sharing links to the lh3.googleusercontent.com form
    Meta can actually fetch. Pass-through for anything that isn't a Drive URL.

    Recognized forms:
      https://drive.google.com/file/d/FILE_ID/view?...
      https://drive.google.com/open?id=FILE_ID
      https://drive.google.com/uc?id=FILE_ID&...
      https://lh3.googleusercontent.com/d/FILE_ID  (already converted, no-op)
    """
    if not url:
        return url
    m = re.search(r"drive\.google\.com/file/d/([A-Za-z0-9_-]+)", url)
    if m:
        return f"https://lh3.googleusercontent.com/d/{m.group(1)}"
    m = re.search(r"drive\.google\.com/(?:open|uc)\?(?:.*&)?id=([A-Za-z0-9_-]+)", url)
    if m:
        return f"https://lh3.googleusercontent.com/d/{m.group(1)}"
    return url


def _split_variants(value):
    """Pipe-separated text becomes a list of variants. Single value stays single."""
    if not value:
        return []
    return [v.strip() for v in value.split("|") if v.strip()]


def _is_multivariant(row):
    for k in ("primary_text", "headline", "description"):
        if "|" in (row.get(k) or ""):
            return True
    return False


def _upload_image(account, url, dry_run):
    """Upload an image to /act_{id}/adimages and return its hash. Required
    for asset_feed_spec (multi-variant) creatives — link_data accepts a raw
    URL via `picture`, but asset_feed_spec only takes hashes."""
    if dry_run:
        return f"DRY_HASH_{abs(hash(url)) % 10**10}"
    import tempfile
    from facebook_business.adobjects.adimage import AdImage

    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = resp.read()
    # SDK's AdImage.remote_create reads the filename field as an on-disk
    # path, not raw bytes. Write to a temp file and point at it.
    tmp = tempfile.NamedTemporaryFile(suffix=".jpg", delete=False)
    try:
        tmp.write(data)
        tmp.close()
        image = AdImage(parent_id=account.get_id_assured())
        image[AdImage.Field.filename] = tmp.name
        image.remote_create()
        return image[AdImage.Field.hash]
    finally:
        os.unlink(tmp.name)


def _drive_video_download_url(url):
    """Drive's lh3.googleusercontent.com/d/<ID> serves image thumbnails,
    not video files. For videos we need the explicit download endpoint."""
    m = re.search(r"drive\.google\.com/file/d/([A-Za-z0-9_-]+)", url)
    if m:
        return f"https://drive.google.com/uc?export=download&id={m.group(1)}"
    m = re.search(r"drive\.google\.com/(?:open|uc)\?(?:.*&)?id=([A-Za-z0-9_-]+)", url)
    if m:
        return f"https://drive.google.com/uc?export=download&id={m.group(1)}"
    m = re.search(r"lh3\.googleusercontent\.com/d/([A-Za-z0-9_-]+)", url)
    if m:
        return f"https://drive.google.com/uc?export=download&id={m.group(1)}"
    return url


def _upload_video(account, url, dry_run):
    """Upload a video to /act_{id}/advideos and return the video_id.
    First tries the file_url path (Meta fetches the URL itself, faster);
    on failure (e.g. Meta can't reach the URL — common with Google Drive)
    falls back to downloading the bytes locally and uploading via the SDK."""
    if dry_run:
        return f"DRY_VIDEO_{abs(hash(url)) % 10**10}"

    # Drive URLs need the explicit download endpoint, not the thumbnail one.
    direct_url = _drive_video_download_url(url)

    try:
        result = account.create_ad_video(params={"file_url": direct_url})
        return result["id"]
    except Exception as exc:
        print(f"  Meta couldn't fetch video URL directly ({exc}). Downloading locally and re-uploading...")

    import tempfile
    from facebook_business.adobjects.advideo import AdVideo

    req = urllib.request.Request(direct_url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=600) as resp:
        data = resp.read()
    # SDK's video upload expects a filepath; write to a temp file.
    tmp = tempfile.NamedTemporaryFile(suffix=".mp4", delete=False)
    try:
        tmp.write(data)
        tmp.close()
        video = AdVideo(parent_id=account.get_id_assured())
        video[AdVideo.Field.filepath] = tmp.name
        video.remote_create()
        return video[AdVideo.Field.id]
    finally:
        os.unlink(tmp.name)


def _video_thumbnail(video_id, dry_run):
    """Get an auto-generated thumbnail URL for an uploaded video. Meta
    requires every video creative to carry image_hash or image_url in
    video_data — even when the video object already has thumbnails — so we
    fetch one and pass it through when the user didn't supply image_url."""
    if dry_run:
        return f"https://dry.thumbnail/{video_id}.jpg"
    from facebook_business.adobjects.advideo import AdVideo
    import time

    # Thumbnails take a few seconds to generate after upload, so poll briefly.
    for _ in range(6):
        thumbs = AdVideo(video_id).get_thumbnails(fields=["uri", "is_preferred"])
        thumbs = list(thumbs)
        if thumbs:
            for t in thumbs:
                if t.get("is_preferred"):
                    return t["uri"]
            return thumbs[0]["uri"]
        time.sleep(2)
    sys.exit(
        f"Video {video_id} has no thumbnail yet — Meta hasn't finished processing it. "
        "Either wait ~1 minute and re-run, or set image_url manually as the thumbnail."
    )


def _wait_for_video_ready(video_id, dry_run, max_wait_seconds=180):
    """Poll a video's processing status until Meta marks it ready, so the
    subsequent /adcreatives call doesn't fail with subcode 1885252 'Video
    not ready for use in an ad' (is_transient=true)."""
    if dry_run:
        return
    from facebook_business.adobjects.advideo import AdVideo
    import time

    deadline = time.time() + max_wait_seconds
    while time.time() < deadline:
        v = AdVideo(video_id).api_get(fields=["status"])
        status = (v.get("status") or {}).get("video_status")
        if status == "ready":
            return
        if status in ("error", "expired"):
            sys.exit(f"Video {video_id} failed processing on Meta's side (status={status!r}).")
        time.sleep(5)
    sys.exit(
        f"Video {video_id} still not ready after {max_wait_seconds}s — Meta is slow processing it. "
        "Re-run the script in a minute or two."
    )


def _build_cta(row):
    """Resolve the call_to_action object from the cta + browser_addon columns.
    browser_addon, when set to anything other than blank/NONE, overrides cta."""
    cta = (row.get("cta") or "").strip()
    link_url = (row.get("link_url") or "").strip()
    addon = (row.get("browser_addon") or "").strip().upper()
    phone = (row.get("phone_number") or "").strip()
    page_id = (row.get("page_id") or "").strip()
    if addon in ("", "NONE"):
        if not link_url and cta != "NO_BUTTON":
            sys.exit(f"Ad {row.get('ad_name')!r}: link_url is empty. Fill in the link_url column.")
        return {"type": cta, "value": {"link": link_url}}
    if addon == "CALL":
        if not phone:
            sys.exit(f"Ad {row.get('ad_name')!r}: browser_addon=CALL needs phone_number.")
        return {"type": "CALL_NOW", "value": {"link": f"tel:{phone}"}}
    if addon == "WHATSAPP":
        if not phone:
            sys.exit(f"Ad {row.get('ad_name')!r}: browser_addon=WHATSAPP needs phone_number (international format, no '+').")
        return {
            "type": "WHATSAPP_MESSAGE",
            "value": {"app_destination": "WHATSAPP", "link": f"https://wa.me/{phone}"},
        }
    if addon == "MESSENGER":
        return {
            "type": "MESSAGE_PAGE",
            "value": {"app_destination": "MESSENGER", "link": f"https://m.me/{page_id}"},
        }
    sys.exit(f"Ad {row.get('ad_name')!r}: unknown browser_addon={addon!r}.")


def _cell_to_str(v):
    """Excel stores numbers as floats; a 15-digit page_id read back is e.g.
    445963815277238.0, which Meta rejects. Strip the .0 for whole-number floats."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _extract_id(value):
    """If the cell holds a label and an ID in parentheses — e.g.
    'Page ABC (123456789)' — return just the ID. Otherwise return the
    trimmed value unchanged."""
    if not value:
        return value
    s = str(value).strip()
    m = re.search(r"\(([^)]+)\)\s*$", s)
    return m.group(1).strip() if m else s


def _normalize_ids(row):
    """Apply _extract_id to every ID-bearing column on the row, including
    comma-separated lists. Mutates row in place."""
    from template_options import ID_COLUMNS, COMMA_SEPARATED_ID_COLUMNS

    for col in ID_COLUMNS:
        if col in row:
            row[col] = _extract_id(row[col])
    for col in COMMA_SEPARATED_ID_COLUMNS:
        if row.get(col):
            row[col] = ",".join(_extract_id(p) for p in row[col].split(",") if p.strip())


# Google Sheets / Excel error values that look like real text but mean
# "this cell's formula broke". Treat any of these as a fatal data error
# rather than passing them through to Meta.
_SHEET_ERROR_VALUES = {
    "#N/A", "#N/A!", "#REF", "#REF!", "#NULL", "#NULL!", "#DIV/0", "#DIV/0!",
    "#VALUE", "#VALUE!", "#NAME", "#NAME?", "#NUM", "#NUM!",
    "#ERROR", "#ERROR!", "#ERROR?", "#GETTING_DATA",
    "#SPILL!", "#CALC!", "#FIELD!", "#CONNECT!", "#UNKNOWN!",
}


def _filter_rows(rows):
    """Drop rows with broken-formula cells (#N/A, #REF!, ...) or an
    empty campaign_name, printing a warning per skipped row. The valid
    rows continue to upload. Exits only if nothing valid is left."""
    valid = []
    for i, row in enumerate(rows, start=2):  # row 1 is the header
        skip_reason = None
        for col, val in row.items():
            if isinstance(val, str) and val.strip().upper() in _SHEET_ERROR_VALUES:
                skip_reason = f"column {col!r} has {val.strip()!r} (broken spreadsheet formula)"
                break
        if not skip_reason and not (row.get("campaign_name") or "").strip():
            skip_reason = "campaign_name is empty"
        if skip_reason:
            print(f"  Skipping row {i}: {skip_reason}")
            continue
        valid.append(row)
    if not valid:
        sys.exit("No valid rows left after skipping broken ones.")
    if len(valid) != len(rows):
        print(f"  Proceeding with {len(valid)} of {len(rows)} row(s).")
    return valid


def _parse_sheet_url(url_or_id):
    """Extract (sheet_id, gid, range) from a Google Sheets URL. Falls
    back to treating the input as a bare sheet_id with gid=0 and the
    default range if it doesn't look like a URL.

    Supports a custom range via `&range=A1:CB200` appended to the URL —
    when present the export endpoint returns only that block instead of
    evaluating the full workbook, which dramatically speeds up fetches
    on large or formula-heavy sheets."""
    m = re.search(r"/spreadsheets/d/([A-Za-z0-9_-]+)", url_or_id)
    sheet_id = m.group(1) if m else url_or_id.strip()
    gid_m = re.search(r"[?&#]gid=(\d+)", url_or_id)
    gid = gid_m.group(1) if gid_m else "0"
    range_m = re.search(r"[?&]range=([A-Za-z0-9:!_\.]+)", url_or_id)
    cell_range = range_m.group(1) if range_m else None
    return sheet_id, gid, cell_range


def load_rows_from_sheet(url_or_id):
    """Fetch a Google Sheet as CSV and parse it. Sheet must be shared
    with 'Anyone with the link → Viewer' so the public export endpoint
    can read it without OAuth."""
    import io

    sheet_id, gid, cell_range = _parse_sheet_url(url_or_id)
    # Default range trims the export to the template's column count
    # (~80 cols → 'CB') and a generous row count, so Google doesn't
    # re-evaluate the whole workbook. Override via `&range=...` in the
    # URL if you have more rows or want a different region.
    if not cell_range:
        cell_range = "A1:CB5000"
    csv_url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/export"
        f"?format=csv&gid={gid}&range={cell_range}"
    )
    req = urllib.request.Request(csv_url, headers={"User-Agent": "Mozilla/5.0"})
    # 10 minutes — generous enough for very large or formula-heavy sheets.
    # The Codespaces / GitHub-app HTTPS proxy still kills browser
    # connections after ~60s though, so for long fetches use the CLI
    # (no proxy in front of it).
    sheet_timeout = int(os.environ.get("META_SHEET_TIMEOUT", "600"))
    try:
        with urllib.request.urlopen(req, timeout=sheet_timeout) as resp:
            ctype = resp.headers.get("Content-Type", "")
            data = resp.read()
    except urllib.error.HTTPError as exc:
        sys.exit(
            f"Failed to fetch Google Sheet (HTTP {exc.code}). The sheet must be "
            "shared with 'Anyone with the link → Viewer' so this script can read "
            "it without OAuth. Open the sheet, click 'Share' top-right, switch "
            "access to that, then retry."
        )
    except (TimeoutError, urllib.error.URLError) as exc:
        sys.exit(
            f"Timed out fetching the Google Sheet after {sheet_timeout}s ({exc}). "
            "Try a narrower range in the URL (e.g. add &range=A1:CB500), or use "
            "the CLI which doesn't have a browser proxy in front of it. Set "
            "META_SHEET_TIMEOUT=900 (or higher) to wait even longer."
        )
    if "text/csv" not in ctype.lower():
        sys.exit(
            f"Google Sheets returned {ctype!r} instead of CSV — the sheet "
            "probably isn't publicly shared. Click 'Share' top-right and switch "
            "access to 'Anyone with the link → Viewer'."
        )
    rows = list(csv.DictReader(io.StringIO(data.decode("utf-8-sig"))))
    if not rows:
        sys.exit(f"Google Sheet {sheet_id} (gid={gid}) has no data rows.")
    for row in rows:
        _normalize_ids(row)
    rows = _filter_rows(rows)
    return rows


def load_rows(path):
    if path.lower().endswith(".xlsx"):
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(it)]
        rows = []
        for raw in it:
            if all(v is None or v == "" for v in raw):
                continue
            rows.append({h: _cell_to_str(v) for h, v in zip(headers, raw)})
    else:
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
    if not rows:
        sys.exit(f"No rows in {path}")
    for row in rows:
        _normalize_ids(row)
    rows = _filter_rows(rows)
    return rows


def group(rows):
    tree = defaultdict(lambda: defaultdict(list))
    campaign_meta = {}
    adset_meta = {}
    for r in rows:
        c, a = r["campaign_name"], r["adset_name"]
        tree[c][a].append(r)
        campaign_meta.setdefault(c, r)
        adset_meta.setdefault((c, a), r)
    return tree, campaign_meta, adset_meta


def special_ad_categories(value):
    if not value or value.strip().upper() == "NONE":
        return []
    return [v.strip() for v in value.split(",") if v.strip()]


def build_targeting(row, dry_run=False):
    saved_id = (row.get("saved_audience_id") or "").strip()
    if saved_id:
        if dry_run:
            return {"saved_audience_id": saved_id}
        from facebook_business.adobjects.savedaudience import SavedAudience

        sa = SavedAudience(saved_id).api_get(fields=["targeting"])
        targeting = sa.get("targeting")
        if not targeting:
            sys.exit(f"Saved audience {saved_id} has no targeting spec — open it in Ads Manager and check it has a location.")
        return targeting
    targeting = {
        "geo_locations": {"countries": [c.strip() for c in row["countries"].split(",") if c.strip()]},
        "age_min": int(row["age_min"]),
        "age_max": int(row["age_max"]),
    }
    genders = _get(row, "genders")
    if genders:
        targeting["genders"] = [int(g.strip()) for g in genders.split(",") if g.strip()]
    incl = _get(row, "included_custom_audience_ids")
    if incl:
        targeting["custom_audiences"] = [{"id": x.strip()} for x in incl.split(",") if x.strip()]
    excl = _get(row, "excluded_custom_audience_ids")
    if excl:
        targeting["excluded_custom_audiences"] = [{"id": x.strip()} for x in excl.split(",") if x.strip()]
    return targeting


def _get(row, key, default=""):
    return (row.get(key) or default).strip() if isinstance(row.get(key), str) else (row.get(key) or default)


def _is_cbo(campaign_row):
    return bool(_money(campaign_row, "campaign_daily_budget") or _money(campaign_row, "campaign_lifetime_budget"))


def build_campaign_params(row, name, multiplier=100):
    params = {
        Campaign.Field.name: name,
        Campaign.Field.objective: row["campaign_objective"],
        Campaign.Field.status: PAUSED,
        Campaign.Field.special_ad_categories: special_ad_categories(row.get("special_ad_categories", "")),
        "is_adset_budget_sharing_enabled": False,
    }
    buying = _get(row, "buying_type")
    if buying:
        params[Campaign.Field.buying_type] = buying

    daily = _money(row, "campaign_daily_budget")
    lifetime = _money(row, "campaign_lifetime_budget")
    if daily and lifetime:
        sys.exit(f"Campaign {name!r}: set campaign_daily_budget OR campaign_lifetime_budget, not both.")
    if daily:
        params["daily_budget"] = int(_parse_amount(daily) * multiplier)
    if lifetime:
        params["lifetime_budget"] = int(_parse_amount(lifetime) * multiplier)

    bid = _get(row, "campaign_bid_strategy")
    if bid:
        params["bid_strategy"] = bid
    cap = _money(row, "campaign_spend_cap")
    if cap:
        params["spend_cap"] = int(_parse_amount(cap) * multiplier)
    start = _get(row, "campaign_start_time")
    if start:
        params["start_time"] = start
    stop = _get(row, "campaign_stop_time")
    if stop:
        params["stop_time"] = stop
    return params


def build_adset_params(row, name, campaign_id, dry_run, campaign_row=None, existing_campaign=False, multiplier=100, is_dynamic=False):
    cbo = _is_cbo(campaign_row) if campaign_row else False
    daily_budget = _money(row, "daily_budget")
    lifetime_budget = _money(row, "lifetime_budget")
    if cbo and (daily_budget or lifetime_budget):
        sys.exit(
            f"Ad set {name!r}: parent campaign uses CBO (campaign_daily_budget or campaign_lifetime_budget set). "
            "Leave daily_budget and lifetime_budget blank on rows under this campaign."
        )
    if not cbo and daily_budget and lifetime_budget:
        sys.exit(f"Ad set {name!r}: set daily_budget OR lifetime_budget, not both.")
    if not existing_campaign and not cbo and not daily_budget and not lifetime_budget:
        sys.exit(f"Ad set {name!r}: needs daily_budget or lifetime_budget (ABO), or a campaign-level budget on the campaign (CBO).")

    bid_strategy = _get(row, "bid_strategy") or "LOWEST_COST_WITHOUT_CAP"
    params = {
        AdSet.Field.name: name,
        AdSet.Field.campaign_id: campaign_id,
        AdSet.Field.billing_event: row["billing_event"],
        AdSet.Field.optimization_goal: row["optimization_goal"],
        AdSet.Field.targeting: build_targeting(row, dry_run=dry_run),
        AdSet.Field.status: PAUSED,
    }
    if not cbo:
        params[AdSet.Field.bid_strategy] = bid_strategy
        if daily_budget:
            params[AdSet.Field.daily_budget] = int(_parse_amount(daily_budget) * multiplier)
        if lifetime_budget:
            params[AdSet.Field.lifetime_budget] = int(_parse_amount(lifetime_budget) * multiplier)

    # bid_amount and bid_roas_floor live on the ad set regardless of CBO/ABO —
    # CBO campaigns still let each ad set carry its own cap / ROAS floor.
    bid_amount = _money(row, "bid_amount")
    if bid_amount:
        params[AdSet.Field.bid_amount] = int(_parse_amount(bid_amount) * multiplier)
    roas_floor = _get(row, "bid_roas_floor")
    if roas_floor:
        # ROAS floor is a percentage (200 = 2.0x ROAS), not currency.
        params["bid_constraints"] = {"roas_average_floor": int(_parse_amount(roas_floor) * 100)}

    daily_cap = _money(row, "daily_spend_cap")
    if daily_cap:
        params["daily_spend_cap"] = int(_parse_amount(daily_cap) * multiplier)
    lifetime_cap = _money(row, "lifetime_spend_cap")
    if lifetime_cap:
        params["lifetime_spend_cap"] = int(_parse_amount(lifetime_cap) * multiplier)
    pacing = _get(row, "pacing_type")
    if pacing:
        params["pacing_type"] = [pacing]

    destination = _get(row, "destination_type")
    if destination:
        params[AdSet.Field.destination_type] = destination
    start = _get(row, "start_time")
    if start:
        params[AdSet.Field.start_time] = start
    end = _get(row, "end_time")
    if end:
        params[AdSet.Field.end_time] = end

    promoted = {}
    pixel_id = _get(row, "pixel_id")
    if pixel_id:
        promoted["pixel_id"] = pixel_id
        event_type = _get(row, "custom_event_type")
        if event_type:
            promoted["custom_event_type"] = event_type
    application_id = _get(row, "application_id")
    if application_id:
        promoted["application_id"] = application_id
    object_store_url = _get(row, "object_store_url")
    if object_store_url:
        promoted["object_store_url"] = object_store_url
    if promoted:
        params[AdSet.Field.promoted_object] = promoted

    dsa_b = _get(row, "dsa_beneficiary")
    if dsa_b:
        params["dsa_beneficiary"] = dsa_b
    dsa_p = _get(row, "dsa_payor")
    if dsa_p:
        params["dsa_payor"] = dsa_p
    if is_dynamic:
        # Required when any creative under this ad set uses
        # asset_feed_spec (multi-variant). Without this Meta rejects
        # /ads with subcode 1885998 'Cannot Create Dynamic Creative
        # ad In Non-Dynamic Creative Ad Set'.
        params["is_dynamic_creative"] = True
    return params


def _apply_advantage_features(spec, row):
    from template_options import ADVANTAGE_FEATURE_COLUMNS, ADVANTAGE_FEATURE_API_SUPPORTED

    def enroll(val):
        v = (val or "").strip().upper()
        return "OPT_IN" if v == "ON" else "OPT_OUT" if v == "OFF" else None

    features_spec = {}
    master = enroll(row.get("advantage_plus_creative"))
    if master:
        # Broad master-switch baseline — applies to both image and video
        # ads regardless of catalog use. Per-feature columns override.
        for f in ("IG_VIDEO_NATIVE_SUBTITLE", "IMAGE_ANIMATION", "TEXT_OVERLAY_TRANSLATION"):
            features_spec[f] = {"enroll_status": master}

    skipped = []
    for col, api_key in ADVANTAGE_FEATURE_COLUMNS:
        per_feature = enroll(row.get(col))
        if not per_feature:
            continue
        if api_key in ADVANTAGE_FEATURE_API_SUPPORTED:
            features_spec[api_key] = {"enroll_status": per_feature}
        else:
            skipped.append(col)
    if skipped:
        print(
            f"  Note: {len(skipped)} adv_* setting(s) skipped (UI-only, "
            f"not exposed in Marketing API): {', '.join(skipped)}. "
            "Configure manually in Ads Manager after upload."
        )
    if features_spec:
        spec["degrees_of_freedom_spec"] = {"creative_features_spec": features_spec}


def build_creative_spec(row, account=None, dry_run=False):
    existing_post = _get(row, "existing_post_id")
    partnership_code = _get(row, "partnership_ad_code")
    if existing_post and partnership_code:
        sys.exit(f"Ad {row.get('ad_name')!r}: set existing_post_id OR partnership_ad_code, not both.")

    # Promote an existing post (your own Page) or run a partner's post via
    # their shared ad code. Both paths use object_story_id and ignore all
    # the creative-content columns (image_url, headline, primary_text, …)
    # because the post already has its content baked in.
    if existing_post or partnership_code:
        if partnership_code:
            if "_" not in partnership_code:
                sys.exit(
                    f"Ad {row.get('ad_name')!r}: partnership_ad_code must be the full "
                    "PARTNER_PAGE_ID_POST_ID format (with underscore) — the partner's "
                    "Page ID is required since it's not your own Page."
                )
            object_story_id = partnership_code
        else:
            object_story_id = existing_post if "_" in existing_post else f"{row['page_id']}_{existing_post}"
        spec = {
            "name": f"Creative - {row['ad_name']}",
            "object_story_id": object_story_id,
        }

        # CTA + link destination override. Meta accepts these on top of an
        # object_story_id creative for most post types — the post's existing
        # CTA/link is replaced for the ad rendering. Skip if NO_BUTTON or
        # either is blank.
        cta = _get(row, "cta")
        link_url = _get(row, "link_url")
        if cta and cta != "NO_BUTTON" and link_url:
            spec["call_to_action"] = {"type": cta, "value": {"link": link_url}}

        # Partnership Ads "Second identity" — Ads Manager's "Second identity"
        # block (Page + IG of the sponsor). Defaults to the row's page_id /
        # instagram_user_id if the dedicated second_identity_* fields are
        # blank.
        if partnership_code:
            second_page = _get(row, "second_identity_page_id") or row.get("page_id")
            if second_page:
                spec["branded_content_sharing_partner_id"] = second_page
            second_ig = _get(row, "second_identity_ig_id") or _get(row, "instagram_user_id")
            if second_ig:
                spec["instagram_user_id"] = second_ig
            # "Identities to display in the header" radio. API field name
            # is a best-guess based on Meta's UI labels.
            display = _get(row, "identity_display").upper()
            if display:
                spec["branded_content_identity_display"] = display

        url_tags = _get(row, "url_tags")
        if url_tags:
            spec["url_tags"] = url_tags
        _apply_advantage_features(spec, row)
        return spec

    video_id = _get(row, "video_id")
    video_url = convert_drive_url(_get(row, "video_url"))
    image_url = convert_drive_url(_get(row, "image_url"))
    if video_id and video_url:
        sys.exit(f"Ad {row.get('ad_name')!r}: set video_id OR video_url, not both.")
    if video_url:
        video_id = _upload_video(account, video_url, dry_run)
    if not video_id and not image_url:
        sys.exit(f"Ad {row.get('ad_name')!r}: needs image_url, video_id, or video_url.")
    cta_obj = _build_cta(row)
    display_link = _get(row, "display_link")

    if _is_multivariant(row):
        spec = _build_asset_feed_creative(row, image_url, video_id, cta_obj, display_link, account, dry_run)
    elif video_id:
        _wait_for_video_ready(video_id, dry_run)
        if not image_url:
            image_url = _video_thumbnail(video_id, dry_run)
        video_data = {
            "video_id": video_id,
            "title": row["headline"],
            "message": row["primary_text"],
            "link_description": row.get("description") or "",
            "call_to_action": cta_obj,
            "image_url": image_url,
        }
        spec = {
            "name": f"Creative - {row['ad_name']}",
            "object_story_spec": {"page_id": row["page_id"], "video_data": video_data},
        }
    else:
        link_data = {
            "link": row["link_url"],
            "message": row["primary_text"],
            "name": row["headline"],
            "description": row["description"],
            "picture": image_url,
            "call_to_action": cta_obj,
        }
        if display_link:
            link_data["caption"] = display_link
        spec = {
            "name": f"Creative - {row['ad_name']}",
            "object_story_spec": {"page_id": row["page_id"], "link_data": link_data},
        }

    instagram_user = _get(row, "instagram_user_id")
    if instagram_user:
        spec["object_story_spec"]["instagram_user_id"] = instagram_user
    threads_user = _get(row, "threads_user_id")
    if threads_user:
        spec["object_story_spec"]["threads_user_id"] = threads_user
    url_tags = _get(row, "url_tags")
    if url_tags:
        spec["url_tags"] = url_tags
    _apply_advantage_features(spec, row)
    return spec


def _build_asset_feed_creative(row, image_url, video_id, cta_obj, display_link, account, dry_run):
    """Multi-variant creative using asset_feed_spec. Triggered when any of
    primary_text / headline / description contains '|'."""
    bodies = _split_variants(row.get("primary_text")) or ([row["primary_text"]] if row.get("primary_text") else [])
    titles = _split_variants(row.get("headline")) or ([row["headline"]] if row.get("headline") else [])
    descriptions = _split_variants(row.get("description")) or ([row["description"]] if row.get("description") else [])

    asset_feed_spec = {
        "bodies": [{"text": t} for t in bodies],
        "titles": [{"text": t} for t in titles],
        "descriptions": [{"text": t} for t in descriptions],
        "link_urls": [{"website_url": row["link_url"], **({"display_url": display_link} if display_link else {})}],
        "call_to_action_types": [cta_obj["type"]],
    }
    if video_id:
        _wait_for_video_ready(video_id, dry_run)
        thumb_url = image_url or _video_thumbnail(video_id, dry_run)
        thumb_hash = _upload_image(account, thumb_url, dry_run)
        asset_feed_spec["videos"] = [{"video_id": video_id, "thumbnail_hash": thumb_hash}]
        asset_feed_spec["ad_formats"] = ["SINGLE_VIDEO"]
    else:
        image_hash = _upload_image(account, image_url, dry_run)
        asset_feed_spec["images"] = [{"hash": image_hash}]
        asset_feed_spec["ad_formats"] = ["SINGLE_IMAGE"]

    return {
        "name": f"Creative - {row['ad_name']}",
        "object_story_spec": {"page_id": row["page_id"]},
        "asset_feed_spec": asset_feed_spec,
    }


def _cleanup(created, account, protected_ids=()):
    """On failure, delete entities we created during this run so the user
    doesn't have to manually clean up orphan campaigns / ad sets / ads /
    creatives.

    `protected_ids` lists entity IDs the user referenced via
    existing_campaign_id / existing_adset_id. They are pre-existing and we
    refuse to delete them under any circumstance, even as a safeguard
    against future bugs in how `created` is populated."""
    protected = set(protected_ids)
    for kind, obj_id in reversed(created):
        if obj_id in protected:
            print(f"  Refusing to delete pre-existing {kind} {obj_id}")
            continue
        try:
            if kind == "ad":
                Ad(obj_id).api_delete()
            elif kind == "adset":
                AdSet(obj_id).api_delete()
            elif kind == "campaign":
                Campaign(obj_id).api_delete()
            elif kind == "creative":
                AdCreative(obj_id).api_delete()
            print(f"  Cleaned up {kind} {obj_id}")
        except Exception as exc:
            print(f"  Failed to clean up {kind} {obj_id}: {exc}")


def _build_campaign_index(account):
    """One-time fetch of all live (non-deleted, non-archived) campaigns in
    the ad account. Returns {name: [campaign_id, ...]} so name-based
    lookups in this run are O(1) and don't hit Meta again."""
    campaigns = account.get_campaigns(
        fields=[Campaign.Field.id, Campaign.Field.name, "effective_status"],
        params={"limit": 1000},
    )
    index = {}
    for c in campaigns:
        if c.get("effective_status") in ("DELETED", "ARCHIVED"):
            continue
        index.setdefault(c["name"], []).append(c["id"])
    return index


def _find_adset_in_campaign(campaign_id, name):
    """Look for an ad set named `name` within an existing campaign.
    Returns the ID if exactly one live match, None if no match. Exits
    with a disambiguation message on 2+ matches."""
    from facebook_business.adobjects.campaign import Campaign as CampaignObj

    adsets = CampaignObj(campaign_id).get_ad_sets(
        fields=[AdSet.Field.id, AdSet.Field.name, "effective_status"],
        params={"limit": 1000},
    )
    matches = [
        a["id"] for a in adsets
        if a["name"] == name and a.get("effective_status") not in ("DELETED", "ARCHIVED")
    ]
    if len(matches) > 1:
        sys.exit(
            f"Found {len(matches)} ad sets named {name!r} in campaign {campaign_id}: "
            f"{', '.join(matches)}. Specify which one in existing_adset_id to disambiguate."
        )
    return matches[0] if matches else None


def upload(account, tree, campaign_meta, adset_meta, dry_run):
    results = []
    created = []
    protected_ids = set()
    # Detect ad account currency so budget/bid amounts entered as
    # "50" become $50 (USD, x100), 50000 won (KRW, x1), etc. Default
    # 100 in dry-run because we don't have a live account to query.
    multiplier = 100
    if account and not dry_run:
        currency = account.api_get(fields=["currency"]).get("currency")
        multiplier = _currency_multiplier(currency)
        print(f"Ad account currency: {currency} — money amounts in template treated as {currency} (x{multiplier} to atomic units)")
    # One-time campaign index for name-based auto-reuse. Empty in dry-run
    # since dry-run shouldn't hit Meta.
    campaign_index = _build_campaign_index(account) if (account and not dry_run) else {}
    try:
        for c_name, adsets in tree.items():
            cm = campaign_meta[c_name]
            existing_campaign = _get(cm, "existing_campaign_id")
            # Auto-lookup by name if existing_campaign_id is blank.
            if not existing_campaign and c_name in campaign_index:
                matches = campaign_index[c_name]
                if len(matches) > 1:
                    sys.exit(
                        f"Found {len(matches)} campaigns named {c_name!r}: "
                        f"{', '.join(matches)}. Specify which one in "
                        "existing_campaign_id to disambiguate."
                    )
                existing_campaign = matches[0]
                print(f"Found existing campaign by name {c_name!r}: {existing_campaign}")
            if existing_campaign:
                campaign_id = existing_campaign
                protected_ids.add(campaign_id)
                print(f"Reusing existing campaign {campaign_id}: {c_name}")
            else:
                c_params = build_campaign_params(cm, c_name, multiplier=multiplier)
                if dry_run:
                    print("CAMPAIGN:", json.dumps(c_params, indent=2))
                    campaign_id = f"DRY_CAMPAIGN_{c_name}"
                else:
                    campaign = account.create_campaign(params=c_params)
                    campaign_id = campaign["id"]
                    created.append(("campaign", campaign_id))
                    print(f"Created campaign {campaign_id}: {c_name}")

            for a_name, ads in adsets.items():
                am = adset_meta[(c_name, a_name)]
                existing_adset = _get(am, "existing_adset_id")
                # An ad set carrying any multi-variant (text-with-|) ad
                # must be created as is_dynamic_creative=True. Reject
                # mixing dynamic + static ads in the same ad set since
                # Meta forbids that.
                dyn_flags = [_is_multivariant(ad) for ad in ads]
                if any(dyn_flags) and not all(dyn_flags):
                    sys.exit(
                        f"Ad set {a_name!r}: contains both ads with text variants "
                        "(using '|' in primary_text/headline/description) and ads "
                        "without. A dynamic creative ad set can only hold dynamic "
                        "ads — split them into two ad sets or remove the variants."
                    )
                is_dynamic = bool(dyn_flags) and all(dyn_flags)
                # Auto-lookup ad set by name within this campaign. Only
                # meaningful when the campaign already exists in Meta (a
                # campaign just created can't have ad sets in it yet).
                if not existing_adset and existing_campaign and not dry_run:
                    matched = _find_adset_in_campaign(campaign_id, a_name)
                    if matched:
                        existing_adset = matched
                        print(f"  Found existing ad set by name {a_name!r}: {existing_adset}")
                if existing_adset:
                    adset_id = existing_adset
                    protected_ids.add(adset_id)
                    print(f"  Reusing existing ad set {adset_id}: {a_name}")
                else:
                    as_params = build_adset_params(am, a_name, campaign_id, dry_run, campaign_row=cm, existing_campaign=bool(existing_campaign), multiplier=multiplier, is_dynamic=is_dynamic)
                    if dry_run:
                        print("AD SET:", json.dumps(as_params, indent=2, default=str))
                        adset_id = f"DRY_ADSET_{a_name}"
                    else:
                        adset = account.create_ad_set(params=as_params)
                        adset_id = adset["id"]
                        created.append(("adset", adset_id))
                        print(f"  Created ad set {adset_id}: {a_name}")

                for ad_row in ads:
                    creative_spec = build_creative_spec(ad_row, account=account, dry_run=dry_run)
                    if dry_run:
                        print("CREATIVE:", json.dumps(creative_spec, indent=2))
                        creative_id = f"DRY_CREATIVE_{ad_row['ad_name']}"
                    else:
                        creative = account.create_ad_creative(params=creative_spec)
                        creative_id = creative["id"]
                        created.append(("creative", creative_id))

                    ad_params = {
                        Ad.Field.name: ad_row["ad_name"],
                        Ad.Field.adset_id: adset_id,
                        Ad.Field.creative: {"creative_id": creative_id},
                        Ad.Field.status: PAUSED,
                    }
                    conversion_domain = _get(ad_row, "conversion_domain")
                    if conversion_domain:
                        ad_params["conversion_domain"] = conversion_domain
                    if dry_run:
                        print("AD:", json.dumps(ad_params, indent=2))
                        ad_id = f"DRY_AD_{ad_row['ad_name']}"
                    else:
                        ad = account.create_ad(params=ad_params)
                        ad_id = ad["id"]
                        created.append(("ad", ad_id))
                        print(f"    Created ad {ad_id}: {ad_row['ad_name']}")

                    results.append({"campaign": campaign_id, "adset": adset_id, "ad": ad_id, "name": ad_row["ad_name"]})
    except Exception:
        if not dry_run and created:
            print("\nUpload failed mid-flight — rolling back created entities:")
            _cleanup(created, account, protected_ids=protected_ids)
        raise
    return results


def main():
    p = argparse.ArgumentParser()
    p.add_argument("source", help="Path to a .csv / .xlsx file, OR a Google Sheets URL (sheet must be shared 'Anyone with the link → Viewer')")
    p.add_argument("--dry-run", action="store_true", help="Print payloads without calling the API")
    args = p.parse_args()

    if args.source.startswith("http") or "docs.google.com" in args.source:
        rows = load_rows_from_sheet(args.source)
    else:
        rows = load_rows(args.source)
    tree, campaign_meta, adset_meta = group(rows)

    account = None
    if not args.dry_run:
        token = os.environ.get("META_ACCESS_TOKEN")
        account_id = os.environ.get("META_AD_ACCOUNT_ID")
        if not token or not account_id:
            sys.exit("Set META_ACCESS_TOKEN and META_AD_ACCOUNT_ID env vars (or use --dry-run).")
        FacebookAdsApi.init(access_token=token)
        account = AdAccount(account_id)

    results = upload(account, tree, campaign_meta, adset_meta, args.dry_run)
    print(f"\nDone. {len(results)} ad(s) staged as PAUSED.")
    if not args.dry_run:
        print("Review in Ads Manager before activating.")


if __name__ == "__main__":
    main()
